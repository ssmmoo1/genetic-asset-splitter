from openpyxl import load_workbook, Workbook
from random import randint
from copy import deepcopy,copy
import tkinter as tk
from tkinter import filedialog
from threading import Thread
import time

filename=""
done = False
newWB = Workbook()
firstFitness = 0
wb = ""
numDone = 0
def splitAssets(filename, popSize,mutationRate,generations,solutions):

    global wb
    wb = load_workbook(filename)
    ws = wb["Sheet1"]
    assets=[]



    totalFiscalValue = 0
    person1EmoValue = 0
    person2EmoValue = 0

    popSize=popSize
    mutationRate = mutationRate
    generations=generations

    class Asset:

        def __init__(self, asset, monetary, emo1, emo2):
            self.asset = asset
            self.monetary = monetary
            self.emo1 = emo1
            self.emo2 = emo2



    class Split():

        def __init__(self,list1,list2):
            self.list1 = list1
            self.list2 = list2
            self.getFitness()
            #print(self.fitness)



        def getFitness(self):

            list1FiscalValue = 0
            list2FiscalValue = 0
            list1emoValue = 0
            list2emoValue = 0

            for items in self.list1:
                list1FiscalValue += items.monetary
                list1emoValue += items.emo1


            for items in self.list2:
                list2FiscalValue += items.monetary
                list2emoValue += items.emo2

            smallestMoney = min(list1FiscalValue,list2FiscalValue)

            fiscalScore = smallestMoney / totalFiscalValue
            fiscalScore = fiscalScore / .5

            smallestEmo = 0
            if(list1emoValue >= list2emoValue):
                smallestEmo = list2emoValue / person2EmoValue
            else:
                smallestEmo = list1emoValue / person1EmoValue

            emoDif = abs(  list1emoValue/person1EmoValue - list2emoValue / person2EmoValue    )
            emoAvg = (list1emoValue / person1EmoValue + list2emoValue / person2EmoValue) /2

            #self.fitness = (  (list1emoValue/person1EmoValue + list2emoValue / person2EmoValue) / 2) - emoDif * 20
            #self.fitness = min( list1emoValue / person1EmoValue, list2emoValue/ person2EmoValue   )

            self.fitness = fiscalScore * 1.2 + smallestEmo - emoDif * 3



        """
        def getFitness(self):
            list1FiscalValue = 0
            list2FiscalValue = 0
            list1emoValue = 0
            list2emoValue = 0
            for items in self.list1:
                list1FiscalValue += items.monetary
                list1emoValue += items.emo1
                list2emoValue += items.emo2
            for items in self.list2:
                list2FiscalValue += items.monetary
            list1FiscalPercent = list1FiscalValue / totalFiscalValue
            list2FiscalPercent = list2FiscalValue / totalFiscalValue
            emo1percent = list1emoValue / person1EmoValue
            emo2percent = list2emoValue / person2EmoValue
            self.fitness = min(list1FiscalPercent, list2FiscalPercent, emo1percent, emo2percent)
        """





    def getNumberOfRows(ws, col): #Does not include the first row containing the titles.
        ws
        nonEmpty = 0
        i = 2
        while (not ws.cell(row=i,column=col).value == None):
            nonEmpty+=1
            i+=1
        return nonEmpty

    def makeFirstPopulation():
        population = [] #will hold split objects
        numAssets = len(assets) #total number from the spreadsheet

        for x in range(popSize):

            list1len = randint( int(numAssets * .2), int(numAssets * .8)  )   #list 1 will have 20-80 percent of assets
            list2len = numAssets - list1len                 #list 2 is the other part

            list1 = []      #two lists to make the split object
            list2 = []

            copyAssets = deepcopy(assets)      #make a copy

            for y in range(list1len):       #fill in list 1 with random assets
                index = randint(0,len(copyAssets)-1)
                list1.append(copyAssets[index])
                del copyAssets[index]

            list2 = copyAssets      # set list 2 to the left overs

            split = Split(list1,list2)  #make the split object
            population.append(split)

        print (population)
        return population  #return an array of split objects
    '''
    def breed(split1, split2):
        lenChoices = []
        lenChoices.append(len(split1.list1))
        lenChoices.append(len(split1.list2))
        lenChoices.append(len(split2.list1))
        lenChoices.append(len(split2.list2))
        childLen = lenChoices[randint(0,3)]
        lenMutate = randint(0,100)
        if(lenMutate < mutationRate):
            childLen += randint(int(childLen* -.1 -1), int(childLen * .1 + 1) )
        assetChoices = []
        assetChoices.append(split1.list1)
        assetChoices.append(split1.list2)
        r = randint(0,1)
        choice1 = assetChoices[r]
        childList1 = []
        childList2 = []
        for x in range(0,len(choice1)):
            if(x < childLen):
                childList1.append(choice1[x])
            else:
                childList2.append(choice1[x])
        copyAssets = deepcopy(assets)
        usedAssets = []
        for x in childList1:
            usedAssets.append(x.asset)
        for x in childList2:
            usedAssets.append(x.asset)
        for items in copyAssets:
            if(not items.asset in usedAssets):
                childList2.append(items)
        assetMutate = randint(0,100)
        if(assetMutate < mutationRate):
            index1 = randint(0,len(childList1)-1)
            index2 = randint(0,len(childList2)-1)
            temp = childList1[index1]
            childList1[index1] = childList2[index2]
            childList2[index2] = temp
        child = Split(childList1,childList2)
        return child
    '''

    def breed(split1, split2):

        child1Len = ( len(split1.list1) + len(split2.list1) ) // 2
        child2Len = len(assets) - child1Len

        assetChoices = deepcopy(split1.list1)
        usedNames = []
        for items in assetChoices:
            usedNames.append(items.asset)


        for items in split2.list1:
            if(not items.asset in usedNames):
                assetChoices.append(items)







        child1 = []
        child2 = []

        for x in range(child1Len):
            index = randint(0,len(assetChoices)-1)
            child1.append( assetChoices[index]  )
            del assetChoices[index]
        copyAssets = deepcopy(assets)

        usedAssets = []
        for x in child1:
            usedAssets.append(x.asset)
        for items in copyAssets:
            if (not items.asset in usedAssets):
                child2.append(items)


        assetMutate = randint(0, 100)
        if (assetMutate < mutationRate):
            index1 = randint(0, len(child1) - 1)
            index2 = randint(0, len(child2) - 1)

            temp = child1[index1]
            child1[index1] = child2[index2]
            child2[index2] = temp





        return Split(child1, child2)

    def makeNextPopulation(oldPop):
        #print('test')

        breedpool = []
        #print("******" + str(len(oldPop)))
        for x in range(0, popSize // 4):
            for y in range(0, int(oldPop[x].fitness * 25)):
                breedpool.append(oldPop[x])

        #print("--------" + str(len(breedpool)))
        newPop = []

        for x in range(0,popSize):
            newPop.append (  breed(breedpool[randint(0,len(breedpool)-1)], breedpool[randint(0,len(breedpool)-1)])  )

        return newPop


    def makeSheet(fittest,sheetNum):
        assets1 = fittest.list1
        assets2 = fittest.list2
        global wb

        wb.create_sheet("Solution " + str(sheetNum))

        ws1 = wb["Solution " + str(sheetNum)]

        ws1.cell(row=1, column=1).value = "Person1's Suggested Assets"
        ws1.cell(row=1, column=2).value = "Person1's Asset Monetary Value"
        ws1.cell(row=1, column=3).value = "Person1's Asset Emotional Value"

        ws1.cell(row=1, column=4).value = "Person2's Suggested Assets"
        ws1.cell(row=1, column=5).value = "Person2's Asset Monetary Value"
        ws1.cell(row=1, column=6).value = "Person2's Asset Emotional Value"

        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 30
        ws1.column_dimensions["C"].width = 30
        ws1.column_dimensions["D"].width = 30
        ws1.column_dimensions["E"].width = 30
        ws1.column_dimensions["F"].width = 30


        for items in assets1:
            nextRow = getNumberOfRows(ws1,1) + 2

            ws1.cell(row=nextRow, column=1).value = items.asset
            ws1.cell(row=nextRow, column=2).value = items.monetary
            ws1.cell(row=nextRow, column=3).value = items.emo1



        for items in assets2:
            nextRow = getNumberOfRows(ws1, 4) + 2

            ws1.cell(row=nextRow, column=4).value = items.asset
            ws1.cell(row=nextRow, column=5).value = items.monetary
            ws1.cell(row=nextRow, column=6).value = items.emo2


        #Calculate Totals
        nextRow1 = getNumberOfRows(ws1,1) + 3
        nextRow2 = getNumberOfRows(ws1,4) + 3

        if(nextRow1 >= nextRow2):
            nextRow = nextRow1
        else:
            nextRow = nextRow2



        ws1.cell(row=nextRow,column = 1).value = "Total "

        ws1.cell(row=nextRow,column = 2).value = "=SUM(B2:B" + str(nextRow-1) + ")"     #monetary
        ws1.cell(row=nextRow, column=5).value = "=SUM(E2:E" + str(nextRow - 1) + ")"

        ws1.cell(row=nextRow, column=3).value = "=SUM(C2:C" + str(nextRow - 1) + ")"        #Emo
        ws1.cell(row=nextRow, column=6).value = "=SUM(F2:F" + str(nextRow - 1) + ")"

        ws1.cell(row=nextRow + 1,column = 1).value = "Monetary and Happiness Percent"

        ws1.cell(row=nextRow + 1,column = 2).value = "=SUM(B2:B" + str(nextRow-1) + ") / " + str(totalFiscalValue)
        ws1.cell(row=nextRow + 1, column=2).number_format = "0.00%"

        ws1.cell(row=nextRow + 1,column = 5).value = "=SUM(E2:E" + str(nextRow-1) + ") / " + str(totalFiscalValue)
        ws1.cell(row=nextRow + 1, column=5).number_format = "0.00%"

        ws1.cell(row=nextRow + 1, column=3).value = "=SUM(C2:C" + str(nextRow - 1) + ") / " + str(person1EmoValue)
        ws1.cell(row=nextRow + 1, column=3).number_format = "0.00%"

        ws1.cell(row=nextRow + 1, column=6).value = "=SUM(F2:F" + str(nextRow - 1) + ") / " + str(person2EmoValue)
        ws1.cell(row=nextRow + 1, column=6).number_format = "0.00%"









    for x in range(2, getNumberOfRows(ws,1) + 2):           #Initialize the asset list
        asset = Asset( ws.cell(row=x,column=1).value, ws.cell(row=x,column=2).value, ws.cell(row=x,column=3).value, ws.cell(row=x,column=4).value)
        assets.append(asset)
        totalFiscalValue += ws.cell(row=x,column=2).value
        person1EmoValue += ws.cell(row=x,column=3).value
        person2EmoValue += ws.cell(row=x,column=4).value


    for y in range(0,solutions):
        global numDone
        #make initial population
        pop = makeFirstPopulation()
        pop.sort(key=lambda x:x.fitness,reverse=True)
        firstFitness = pop[0].fitness
        print("First highest fitness " + str(firstFitness))


        for x in range(generations):
            #makeSheet(pop[0])
            pop = makeNextPopulation(pop)
            pop.sort(key=lambda x:x.fitness,reverse=True)

        for x in range(len(pop)):
            print(pop[x].fitness)

        print( pop[0].fitness )

        makeSheet(pop[0],y)
        numDone +=1

    saveFile = filedialog.asksaveasfilename(initialdir="/", title="Select File", filetypes=(("Excel Files", "*.xlsx"), ("all files", "*.*")))











    #newWB.remove_sheet("Sheet")
    print(wb.get_sheet_names())
    #newWB.remove(newWB.get_sheet_by_name("Sheet"))
    wb.save(saveFile + ".xlsx")
    global done
    done = True
    exit()

root = tk.Tk()
root.title("Asset Splitter")
firstFrame = tk.Frame(master=root)
popSize = 150
mutationRate=5
generations=75
solutions = 10

filename = "assets.xlsx"


def checkDone():
    if done == True:
        exit()
    else:
        root.after(10,func=checkDone)

def chooseFile():
    global filename
    filename = filedialog.askopenfilename(initialdir="/", title="Select File", filetypes=(("Excel Files", "*.xlsx"), ("all files", "*.*")))

def startBut():
    global popSize
    global mutationRate
    global generations
    global filename
    global solutions


    solutions = int(solutionsEntry.get())

    secondFrame = tk.Frame(master=root)
    label = tk.Label(secondFrame,text="Assets are splittings please \nwait for the save window to pop up. \n This may take several minutes.\n",font=("Helvetica",30))
    label.pack()
    secondFrame.grid(row=0,column=0)
    secondFrame.tkraise()
    root.update_idletasks()
    root.update()
    t1 = Thread(target=splitAssets,args=[filename,popSize,mutationRate,generations,solutions])
    t1.start()

    while numDone != solutions:
        print('checking if done')
        time.sleep(.2)
        label.config(text="Assets are splittings please \nwait for the save window to pop up. \n This may take several minutes.\n Completed: " + str(numDone) + "/" + str(solutions),font=("Helvetica",30))
        root.update_idletasks()
        root.update()


firstFrame.grid(row=0,column=0)


solutionsLabel = tk.Label(firstFrame,text="Enter number of solutions")



solutionsEntry = tk.Entry(firstFrame)
solutionsEntry.insert(0,solutions)

fileButton = tk.Button(firstFrame,command=chooseFile,text = "Choose File")
startButton = tk.Button(firstFrame,command=startBut, text="Start")

instructionsLabel = tk.Label(firstFrame, text="First click choose file to select an excel file \n Then choose the number of solutions you want the program to generate \n Click start and wait to save the new file. \n It may take several minutes to complete")

instructionsLabel.grid(row=0, column=1)

solutionsLabel.grid(row=2,column=1)


solutionsEntry.grid(row=3,column=1)

fileButton.grid(row=1,column=1)
startButton.grid(row=4,column=1)









#label2 = tk.Label(firstFrame, text="Assets are Splitting. Please wait to save the split assets.",font=("Helvetica",20))
#label2.pack()


root.after(10,func=checkDone)
root.mainloop()

#read from spreadsheet and make the asset objects
